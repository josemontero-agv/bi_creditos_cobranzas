from flask import Blueprint, render_template, jsonify, request, Response
from flask_login import login_required
from ..services.odoo_connector import OdooConnector
from ..services.kpi_calculator import compute_kpis, top15_clients


main_bp = Blueprint("main", __name__, template_folder="templates")


@main_bp.route("/")
def index():
    return render_template("dashboard.html")


@main_bp.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")


@main_bp.route("/reports")
@login_required
def reports_view():
    return render_template("reports.html")


@main_bp.route("/api/kpis")
@login_required
def api_kpis():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    customer = request.args.get('q')
    try:
        connector = OdooConnector()
        invoices = connector.get_unpaid_invoices(start_date=start_date, end_date=end_date, customer=customer)
        kpis = compute_kpis(invoices)
        # Compose extra datasets for dashboard
        # Top 10 clientes por monto vencido
        top_clients_vencido = {}
        cond_map = {"vigente": 0.0, "vencido": 0.0}
        tipo_doc_map = {}
        from datetime import date, datetime, timedelta
        today = date.today()
        # Para serie mensual del índice de morosidad
        month_totals = {}
        month_vencido = {}
        for inv in invoices:
            residual = float(inv.get("amount_residual") or 0.0)
            if residual <= 0:
                continue
            # condición
            due = inv.get("invoice_date_due")
            is_vencido = False
            if due:
                try:
                    is_vencido = datetime.strptime(due, "%Y-%m-%d").date() < today
                except Exception:
                    is_vencido = False
            cond_map["vencido" if is_vencido else "vigente"] += residual
            # top clientes (solo vencido)
            if is_vencido:
                partner = inv.get("partner_id")
                name = partner[1] if isinstance(partner, list) and len(partner) >= 2 else str(partner)
                top_clients_vencido[name] = top_clients_vencido.get(name, 0.0) + residual
            # tipo de documento
            doc = inv.get("l10n_latam_document_type_id")
            doc_name = doc[1] if isinstance(doc, list) and len(doc) >= 2 else (inv.get("move_type") or "Desconocido")
            tipo_doc_map[doc_name] = tipo_doc_map.get(doc_name, 0.0) + residual

            # Serie mensual índice de morosidad
            inv_date = inv.get("invoice_date")
            try:
                month_key = datetime.strptime(inv_date, "%Y-%m-%d").strftime("%Y-%m") if inv_date else None
            except Exception:
                month_key = None
            if month_key:
                month_totals[month_key] = month_totals.get(month_key, 0.0) + residual
                # vencido en el mes si fecha de vencimiento <= último día del mes
                if due:
                    try:
                        due_date = datetime.strptime(due, "%Y-%m-%d").date()
                        # último día del mes
                        year, month = map(int, month_key.split("-"))
                        if month == 12:
                            last_day = date(year, 12, 31)
                        else:
                            first_next = date(year, month + 1, 1)
                            last_day = first_next - timedelta(days=1)
                        if due_date <= last_day:
                            month_vencido[month_key] = month_vencido.get(month_key, 0.0) + residual
                    except Exception:
                        pass

        # Top 10 sorted
        top_sorted = sorted(top_clients_vencido.items(), key=lambda x: x[1], reverse=True)[:10]
        top10_labels = [n for n, _ in top_sorted]
        top10_values = [round(v, 2) for _, v in top_sorted]

        # Serie ordenada por mes
        months_sorted = sorted(month_totals.keys())
        serie_labels = months_sorted
        serie_values = []
        for mk in months_sorted:
            total = month_totals.get(mk, 0.0)
            venc = month_vencido.get(mk, 0.0)
            serie_values.append(round((venc / total * 100) if total else 0.0, 2))

        resp = dict(kpis)
        resp.update({
            "top10": {"labels": top10_labels, "values": top10_values},
            "condicion": {"labels": ["Vigente", "Vencido"], "values": [round(cond_map["vigente"],2), round(cond_map["vencido"],2)]},
            "tipo_documento": {"labels": list(tipo_doc_map.keys()), "values": [round(v,2) for v in tipo_doc_map.values()]},
            "morosidad_series": {"labels": serie_labels, "values": serie_values}
        })
        return jsonify(resp)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@main_bp.route("/api/reports/top15")
@login_required
def api_top15():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    customer = request.args.get('q')
    try:
        connector = OdooConnector()
        invoices = connector.get_unpaid_invoices(start_date=start_date, end_date=end_date, customer=customer)
        clientes, montos = top15_clients(invoices)
        return jsonify({"clientes": clientes, "montos": montos})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@main_bp.route("/api/odoo/ping")
@login_required
def api_odoo_ping():
    try:
        connector = OdooConnector()
        connector.ping()
        return jsonify({"ok": True, "detail": "Conexión a Odoo OK"})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@main_bp.route("/api/public/odoo/ping")
def api_odoo_ping_public():
    try:
        connector = OdooConnector()
        connector.ping()
        return jsonify({"ok": True, "detail": "Conexión a Odoo OK"})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@main_bp.route("/api/reports/top15/details")
@login_required
def api_top15_details():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    customer = request.args.get('q')
    try:
        connector = OdooConnector()
        invoices = connector.get_unpaid_invoices(start_date=start_date, end_date=end_date, customer=customer)
        # Determinar top15 para filtrar
        clientes, _ = top15_clients(invoices)
        clientes_set = set(clientes)
        details = []
        for inv in invoices:
            partner = inv.get("partner_id")
            partner_name = None
            if isinstance(partner, list) and len(partner) >= 2:
                partner_name = str(partner[1])
            elif isinstance(partner, str):
                partner_name = partner
            else:
                partner_name = "(Sin nombre)"
            if partner_name not in clientes_set:
                continue
            details.append({
                "cliente": partner_name,
                "documento": inv.get("name"),
                "fecha": inv.get("invoice_date"),
                "vence": inv.get("invoice_date_due"),
                "monto": float(inv.get("amount_total") or 0.0),
                "saldo": float(inv.get("amount_residual") or 0.0),
                "origen": inv.get("invoice_origin") or "",
            })
        # Ordenar por saldo desc
        details.sort(key=lambda x: x["saldo"], reverse=True)
        return jsonify({"rows": details})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@main_bp.route('/api/reports/top15/details.csv')
@login_required
def api_top15_details_csv():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    customer = request.args.get('q')
    try:
        connector = OdooConnector()
        invoices = connector.get_unpaid_invoices(start_date=start_date, end_date=end_date, customer=customer)
        clientes, _ = top15_clients(invoices)
        clientes_set = set(clientes)
        lines = ["Cliente,Documento,Fecha,Vence,Monto,Saldo,Origen"]
        for inv in invoices:
            partner = inv.get("partner_id")
            partner_name = None
            if isinstance(partner, list) and len(partner) >= 2:
                partner_name = str(partner[1])
            elif isinstance(partner, str):
                partner_name = partner
            else:
                partner_name = "(Sin nombre)"
            if partner_name not in clientes_set:
                continue
            monto = float(inv.get("amount_total") or 0.0)
            saldo = float(inv.get("amount_residual") or 0.0)
            row = [
                partner_name.replace(',', ' '),
                str(inv.get("name") or ''),
                str(inv.get("invoice_date") or ''),
                str(inv.get("invoice_date_due") or ''),
                f"{monto:.2f}",
                f"{saldo:.2f}",
                str(inv.get("invoice_origin") or '').replace(',', ' ')
            ]
            lines.append(','.join(row))
        csv_data = '\n'.join(lines)
        return Response(csv_data, mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=top15_detalle.csv'})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@main_bp.route('/api/reports/data')
@login_required
def api_reports_data():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    customer = request.args.get('q')
    try:
        connector = OdooConnector()
        rows = connector.get_report_lines(start_date=start_date, end_date=end_date, customer=customer)
        return jsonify({"rows": rows})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@main_bp.route('/api/reports/export.xlsx')
@login_required
def api_reports_export_xlsx():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    customer = request.args.get('q')
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        connector = OdooConnector()
        rows = connector.get_report_lines(start_date=start_date, end_date=end_date, customer=customer)
        wb = Workbook()
        ws = wb.active
        ws.title = 'CxC'
        # Map de etiquetas legibles
        columns = [
            ('date','Fecha (date)'),
            ('I10nn_latam_document_type_id','Tipo de Documento (l10n_latam_document_type_id)'),
            ('move_name','Número (move_name)'),
            ('invoice_origin','Origen (invoice_origin)'),
            ('account_id/code','Cuenta/Código (account_id/code)'),
            ('account_id/name','Cuenta/Nombre (account_id/name)'),
            ('patner_id/cod_client_sap','Socio/Cod. Cliente SAP (partner_id/code_client_sap)'),
            ('patner_id/vat','Socio/NIF (partner_id/vat)'),
            ('patner_id','Socio (partner_id)'),
            ('amount_currency','Importe en moneda (amount_currency)'),
            ('amount_residual_currency','Importe residual en moneda (amount_residual_currency)'),
            ('date_maturity','Fecha de vencimiento (date_maturity)'),
            ('ref','Referencia (ref)'),
            ('name','Etiqueta (name)'),
            ('move_id/invoice_user_id','Asiento/Vendedor (move_id/invoice_user_id)'),
            ('patner_id/state_id','Socio/Provincia (partner_id/state_id)'),
            ('patner_id/l10n_pe_district','Socio/Distrito (partner_id/l10n_pe_district)'),
            ('patner_id/contact_adress','Socio/Dirección completa (partner_id/contact_adress)'),
            ('destiny_adress','Dirección de destino (destiny_adress)'),
            ('patner_id/country_code','Socio/Código de país (partner_id/country_code)'),
            ('patner_id/country_id','Socio/País (partner_id/country_id)'),
            ('move_id/sales_channel_id','Asiento/Canal de venta (move_id/sales_channel_id)'),
            ('move_id/sales_type_id','Asiento/Tipo de venta (move_id/sales_type_id)')
        ]
        ws.append([label for _, label in columns])
        for r in rows:
            ws.append([r.get(key, '') for key, _ in columns])
        # Adjust column widths
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(12, len(h) + 2), 40)
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(output.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=cxc_report.xlsx'})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

